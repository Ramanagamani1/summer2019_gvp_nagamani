import click
import openpyxl
import urllib.request
import html.parser
from bs4 import BeautifulSoup
import MySQLdb as MS
global db1,cursor
@click.group()
def main():
    db1=MS.connect(host="localhost",user="root",passwd="daddy")
    cursor=db1.cursor()


@main.command()
def importdata():
    wb=openpyxl.load_workbook('students.xlsx')
    sheet=wb.get_sheet_by_name('Colleges')
    max_row=sheet.max_row
    con=MS.connect('localhost','root','daddy','onlinedb')
    cursor=con.cursor()
    query="insert into onlineapp_college (id,name,location,acronym,contact) values (%s,%s,%s,%s,%s)"
    for r in range(2,max_row+1):
         id=r-1
         name=sheet.cell(r,1).value
         location=sheet.cell(r,3).value
         acronym=sheet.cell(r,2).value
         contact=sheet.cell(r,4).value
         values = (id,name,location,acronym,contact)
         cursor.execute(query,values)
    #import data from html
    page = "https://d1b10bmlvqabco.cloudfront.net/attach/inpg92dp42z2zo/hdff4poirlh7i6/io5hun2sdr21/mock_results.html"
    soup = BeautifulSoup(urllib.request.urlopen(page).read(), 'html.parser')
    title = soup.h1.string
    tbody = soup('table')[0].find_all('tr')
    list1 = []
    for row in tbody:
        cols = row.findChildren(recursive=False)
        cols = [ele.text.strip() for ele in cols]
        list1.append(cols[0:])
    print(list1)
    query1="insert into onlineapp_mocktest1 (id,problem1,problem2,problem3,problem4,total,student_id) values (%s,%s,%s,%s,%s,%s,%s)"
    for r in range(2,len(list1)+1):
        id=list1[r-1][0]
        print(id, end=" ")
        problem1=list1[r-1][2]
        problem2 = list1[r - 1][3]
        problem3 = list1[r - 1][4]
        problem4 = list1[r - 1][5]
        total=list1[r-1][6]
        student_id=list1[r-1][0]
        values=(id,problem1,problem2,problem3,problem4,total,student_id)
        cursor.execute(query1,values)
    con.commit()


if __name__=='__main__':
    main()