import os

import click
import openpyxl
import urllib.request
import html.parser
from bs4 import BeautifulSoup
import MySQLdb as MS
@click.group()
def main():
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
    import django
    django.setup()
    pass

@main.command()
def importdata():
    wb = openpyxl.load_workbook('students.xlsx')
    sheet = wb.get_sheet_by_name('Colleges')
    max_row = sheet.max_row
    lists=[]
    for r in range(2,max_row+1):
         l=[]
         id=r-1
         name=sheet.cell(r,1).value
         location=sheet.cell(r,3).value
         acronym=sheet.cell(r,2).value
         contact=sheet.cell(r,4).value
         l.append(id)
         l.append(name)
         l.append(location)
         l.append(acronym)
         l.append(contact)
         lists.append(l)
    sheet1=wb.get_sheet_by_name('Current')
    max_row1=sheet1.max_row
    list2=[]
    for r in range(2,max_row1+1):
        l=[]
        name=sheet1.cell(r,1).value
        email=sheet1.cell(r,3).value
        college=sheet1.cell(r,2).value
        dbfolder=sheet1.cell(r,4).value
        l.append(name)
        l.append(email)
        l.append(college)
        l.append(dbfolder)
        list2.append(l)

    from onlineapp.models import College, MockTest1, Student
    i=0
    from onlineapp.models import College, MockTest1,Student
    for l in lists:
        c=College(name=lists[i][1],location=lists[i][2],acronym=lists[i][3],contact=lists[i][4])
        i=i+1
        c.save()

    sheet2 = wb.get_sheet_by_name('Deletions')
    max_row2=sheet2.max_row
    list3=[]
    for r in range(2,max_row2+1):
        l=[]
        name=sheet2.cell(r,1).value
        email=sheet2.cell(r,3).value
        college=sheet2.cell(r,2).value
        dbfolder=sheet2.cell(r,4).value
        l.append(name)
        l.append(email)
        l.append(college)
        l.append(dbfolder)
        list3.append(l)
    print(list3)
    i=0
    for l in list2:
        c=Student(name=list2[i][0],dob='1999-11-21',email=list2[i][1],db_folder=list2[i][3],dropped_out=False)
        clg_names=College.objects.all()
        for x in clg_names:
            if x.acronym==list2[i][2]:
                c.college=College.objects.get(id=x.id)
                break
        i = i + 1
        c.save()

    i=0
    for l in list3:
        c=Student(name=list3[i][0],dob='1999-11-21',email=list3[i][1],db_folder=list3[i][3],dropped_out=True)
        clg_names=College.objects.all()
        for x in clg_names:
            if x.acronym==list3[i][2]:
                c.college=College.objects.get(id=x.id)
                break
        i = i + 1
        c.save()


    #extract data from html and insert into MockTest1 table
    page = "https://d1b10bmlvqabco.cloudfront.net/attach/inpg92dp42z2zo/hdff4poirlh7i6/io5hun2sdr21/mock_results.html"
    soup = BeautifulSoup(urllib.request.urlopen(page).read(), 'html.parser')
    title = soup.h1.string
    tbody = soup('table')[0].find_all('tr')
    list1 = []
    for row in tbody:
        cols = row.findChildren(recursive=False)
        cols = [ele.text.strip() for ele in cols]
        list1.append(cols[0:])

    for l in range(1,len(list1)):
        c=MockTest1(problem1=list1[l][2],problem2=list1[l][3],problem3=list1[l][4],problem4=list1[l][5],total=list1[l][6])
        z=list1[l][1]
        z=z.split('_')
        val=z[2]
        st_id=Student.objects.all()
        for i in st_id:
            if i.db_folder.lower()==val:
               try:
                  c.student=Student.objects.get(pk=i.id)
                  c.save()
                  break
               except ('IntegrityError','Error'):
                   continue

    #print(list1)

@main.command()
def delete1():
    from onlineapp.models import College, MockTest1, Student
    c=College.objects.all()
    for r in c:
        r.delete()


if __name__=='__main__':
    main()